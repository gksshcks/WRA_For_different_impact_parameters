import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from openpyxl import load_workbook
from matplotlib.ticker import ScalarFormatter
import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.cm as cm
from openpyxl import load_workbook
from matplotlib.ticker import ScalarFormatter

# Define the file paths for the data representing the infinitesimal Wigner rotation angle (IWRA)
# collected from various locations and under different conditions:
# 1. From a 300km Low Earth Orbit (LEO)
# 2. From a 300km Low Earth Orbit (LEO) for photons with negative impact parameters
# 3. From Earth's surface
# 4. From Earth's surface for photons with negative impact parameters
# 5. From the vicinity of a Black Hole (BH) for an observer with zero Q
# 6. From the vicinity of a Black Hole (BH)
#    for photons with negative impact parameters and an observer with zero Q
# 7. From the vicinity of a Black Hole (BH) with zero angular momentum (J) for an observer with zero Q
# 8. From the vicinity of a Black Hole (BH) with zero angular momentum (J),
#    for photons with negative impact parameters and an observer with zero Q and negative impact parameters


file_path = 'Data/tildepsir_from_300kmLEO.xlsx'
file_path_m = 'Data/tildepsir_from_300kmLEO_negative_impact_parameters.xlsx'

file_path_cofactor_rearth = 'Data/1+n3fromEarth.xlsx'
file_path_cofactor_rearth_m = 'Data/1+n3fromEarthb.xlsx'

file_path_rearth= 'Data/tildepsir_from_Earth.xlsx'
file_path_rearth_m= 'Data/tildepsir_from_Earth_negative_impact_parameters.xlsx'

file_path_BH= 'Data/tildepsi_from_BH_with_zeroQ.xlsx'
file_path_BH_m= 'Data/tildepsi_from_BH_with_zeroQ_negative_impact_parameters.xlsx'
file_path_BH_zeroJ= 'Data/tildepsi_from_BH_with_zeroQ_zeroJ.xlsx'
file_path_BH_m_zeroJ= 'Data/tildepsi_from_BH_zeroQ_zeroJ_negative_impact_parameters.xlsx'

file_path_cofactor_BH = 'Data/1+n3fromBHzeroLz.xlsx'
file_path_cofactor_BH_m = 'Data/1+n3fromBHzeroLzm.xlsx'

file_path_BH_zeroLz = 'Data/1+n3fromBHzeroLz.xlsx'
file_path_BH_zeroLz_m = 'Data/1+n3fromBHzeroLzm.xlsx'

def WRA_function(file_path,negative_params=False, Earth=True, integer_params=True, cofactor=False):
    if Earth:
        r_sender = 6378
        interval = 100
    else:
        r_sender = 83624391900
        interval = (13/200)

    "Load the Excel file using openpyxl and Create an empty dataframe (called data)"
    wb = load_workbook(file_path)
    sheets = wb.sheetnames
    data = pd.DataFrame()

    "Define some constants and parameters for processing the sheets" \
    "Each sheet corresponds to the ratio rk^(phi)/k^r from 0.5 to 5 with step of 0.5"
    number_of_sheets = 10
    step = 0.5
    sheet_to_parameter = {f'Sheet{i + 1}': (i * step) + 0.5 for i in range(number_of_sheets)}

    "Appending the data from each parameter to the empy Dataframe"
    for sheet in sheets:
        if Earth:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, names=['parameter', 'radius', 'infWRA'])
        else:
            df = pd.read_excel(file_path, sheet_name=sheet, header=None, names=['parameter', 'radius', 'affine parameter','infWRA'])
        df['parameter'] = sheet_to_parameter[sheet]
        # For negative parameter, negating the parameters
        if negative_params:
            df['parameter'] = -1 * df['parameter']
        # Append the data from the current sheet to the main DataFrame
        data = data.append(df)

    if cofactor:
        for sheet in sheets:
            if Earth:
                df = pd.read_excel(file_path, sheet_name=sheet, header=None, names=['parameter', 'radius', 'cofactor'])
            else:
                df = pd.read_excel(file_path, sheet_name=sheet, header=None,
                                   names=['parameter', 'radius', 'affine parameter', 'cofactor'])
            df['parameter'] = sheet_to_parameter[sheet]
            # For negative parameter, negating the parameters
            if negative_params:
                df['parameter'] = -1 * df['parameter']
            # Append the data from the current sheet to the main DataFrame
            data = data.append(df)

    "Filtering conditions to select specific data and Apply the filtering condition"
    if integer_params:
        condition = (
                ((data['parameter'] >= -5) & (data['parameter'] < -1) & (data['parameter'].astype(int) == data['parameter'])) |
                ((data['parameter'] > 1) & (data['parameter'] <= 5) & (data['parameter'].astype(int) == data['parameter']))
        )
    else:
        condition = ((data['parameter'] >= -5) & (data['parameter'] < -1)) \
                    | ((data['parameter'] > 1) & (data['parameter'] <= 5))
    data = data[condition]

    if not cofactor:
        "# Calculate the cumulative sum to integrate the 'infWRA' columns." \
        "Here, we multiply by 100 to account for the fact that " \
        "each data point in the dataset is sampled at an interval of 100 km " \
        "in radial distance from the origin (x,y,z)=(0,0,0)."
        data['integrated_infWRA'] = data.groupby('parameter')['infWRA'].cumsum() * interval
    data['radius_minus_r_sender'] = data['radius'] - r_sender  # 6378 is the radius of Earth

    return data
def plot_wigner_rotation(file_path, Earth=True, negative_params=False, ax=None, cofactor=False, custom_ticks=False, show_legend=True):
    data = WRA_function(file_path, Earth=Earth, negative_params = negative_params, cofactor=cofactor)

    # Get unique parameter values and sort them
    parameters = data['parameter'].unique()
    parameters.sort()
    parameters = parameters[::-1]

    # Set up the plot
    plt.style.use('seaborn-notebook')
    if ax is None:
        fig, ax = plt.subplots(figsize=(8, 5))

    # Define colormap for plotting
    if negative_params:
        colormap = cm.Blues(np.linspace(0.3, 1, len(parameters)))
    else:
        colormap = cm.Reds(np.linspace(1, 0.3, len(parameters)))
    if cofactor:
        # Plot data for each parameter with a different color
        for param, color in zip(parameters, colormap):
            param_data = data[data['parameter'] == param]
            label_str = r'$\frac{rk^\phi}{k^r} (\approx b_{ph})$' + f'= {param}'
            ax.plot(param_data['radius_minus_r_sender'], param_data['cofactor'], label=label_str, color=color)
        # Set axis labels, ticks, and legend
        ax.set_xlabel('Altitude (km)', fontsize=18)
        ax.set_ylabel(r'$\frac{1}{1+n^\hat{3}}$', fontsize=18)
    else:
        for param, color in zip(parameters, colormap):
            param_data = data[data['parameter'] == param]
            label_str = r'$\frac{rk^\phi}{k^r} (\approx b_{ph})$' + f'= {param}'
            ax.plot(param_data['radius_minus_r_sender'], param_data['integrated_infWRA'], label=label_str, color=color)
        # Set axis labels, ticks, and legend
        ax.set_xlabel('Altitude (km)', fontsize=18)
        ax.set_ylabel('Wigner rotation angle (degree)', fontsize=18)

    if Earth:
        xmin, xmax = data['radius_minus_r_sender'].min(), data['radius_minus_r_sender'].max()
    else:
        xmax = data.groupby('parameter')['radius_minus_r_sender'].max().min()
    ax.set_xlim(0, xmax)

    for axis in [ax.xaxis, ax.yaxis]:
        formatter = ScalarFormatter(useMathText=True)
        formatter.set_scientific(True)
        formatter.set_powerlimits((-3, 3))
        axis.set_major_formatter(formatter)
        axis.offsetText.set_fontsize(14)

    if Earth:
        step = 10000
        xticks = np.arange(0, xmax, step=step)
        ax.set_xticks(xticks)
        if not cofactor:
            ax.set_yscale('symlog', linthresh=1e-8)



    ax.tick_params(axis='both', which='major', labelsize=15)

    if show_legend:
        # Customize the legend
        ncol_value = 2 if negative_params else 4
        legend = ax.legend(
            ncol=ncol_value,
            loc='center',
            fontsize=13,
            bbox_to_anchor=(1.01, 1) if not negative_params else None
        )

    # Tighten the layout
    plt.tight_layout()
    return ax

def plot_wigner_rotation_for_both_positive_and_negative_impact_parameter(
        file_path_for_positive_impact_factor, file_path_for_negative_impact_factor):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path_for_positive_impact_factor, ax=ax)
    plot_wigner_rotation(file_path_for_negative_impact_factor, negative_params=True, ax=ax)

    plt.savefig('Results/WRA_combined Earth' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/WRA_combined Earth' + '.svg', dpi=300, bbox_inches='tight')

    plt.show()
def plot_cofactor_for_positive_impact_parameter(
        file_path_for_positive_impact_factor):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path_for_positive_impact_factor, ax=ax, cofactor=True, custom_ticks=False, show_legend=False)


    plt.savefig('Results/cofactor_Earth_positive' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/cofactor_Earth_positive' + '.svg', dpi=300, bbox_inches='tight')

    plt.show()
def plot_cofactor_for_negative_impact_parameter_BH(
        file_path_for_negative_impact_factor):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path_for_negative_impact_factor,  Earth=False, negative_params=True, ax=ax, cofactor=True,
                         custom_ticks=False,
                         show_legend=False)
    plt.savefig('Results/cofactor_BH_negative' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/cofactor_BH_negative' + '.svg', dpi=300, bbox_inches='tight')

    plt.show()
def plot_cofactor_for_positive_impact_parameter_BH(
        file_path_for_positive_impact_factor):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path_for_positive_impact_factor, ax=ax,  Earth=False, cofactor=True, custom_ticks=False, show_legend=False)


    plt.savefig('Results/cofactor_BH_positive' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/cofactor_BH_positive' + '.svg', dpi=300, bbox_inches='tight')

    plt.show()
def plot_cofactor_for_negative_impact_parameter(
        file_path_for_negative_impact_factor):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path_for_negative_impact_factor, negative_params=True, ax=ax, cofactor=True,
                         custom_ticks=False,
                         show_legend=False)
    plt.savefig('Results/cofactor_Earth_negative' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/cofactor_Earth_negative' + '.svg', dpi=300, bbox_inches='tight')

    plt.show()
def plot_wigner_rotation_difference(file_path1, file_path2):
    # Load data from the specified file paths and calculate the Wigner Rotation Angle (WRA) difference
    data_positive_impact_factor = WRA_function(file_path1, integer_params=False)
    data_negative_impact_factor = WRA_function(file_path2, integer_params=False)

    data_positive_impact_factor.set_index(['parameter', 'radius', 'radius_minus_r_sender'], inplace=True)
    data_negative_impact_factor.set_index(['parameter', 'radius', 'radius_minus_r_sender'], inplace=True)

    # Calculate the difference in WRA between positive and negative impact paramete
    #, used for caluclating Eq. 27 in Supplementary Information
    data_diff = data_positive_impact_factor['infWRA'] - data_negative_impact_factor['infWRA']
    data_diff = data_diff.reset_index()
    parameters = data_diff['parameter'].unique()

    min_x, max_x = [], []
    min_result = []
    max_result = []
    rmin = 1500

    # Set up the plot
    fig, ax = plt.subplots(figsize=(8, 8))
    colors = cm.magma(np.linspace(0.1, 0.9, len(parameters)))
    plt.style.use('seaborn-notebook')

    for param, color in zip(parameters, colors):
        param_data = data_diff[data_diff['parameter'] == param].copy()
        param_data['radius_David'] = param_data['radius'].copy()
        param_data['radius_David_minus_r_sender'] = param_data['radius_David'] - 6378
        param_data = param_data[param_data['radius_David_minus_r_sender'] >= 0].copy()

        for index, radius_David in enumerate(param_data['radius_David']):
            Sol1 = np.roots([1,
                             -2 * (6378 + 300) * (np.cos(np.pi - np.arctan(param))),
                             (6378 + 300) ** 2 - radius_David ** 2])

            alpha = np.arctan(param)
            beta = np.arccos(
                ((6378 + 300) ** 2 + radius_David ** 2 - Sol1[1] ** 2) / (2 * (6378 + 300) * radius_David)) / 2
            h = (param_data['radius_David_minus_r_sender']) / (np.tan(np.pi / 2 - alpha + 2 * beta) + 1 / param)
            radius_condition = np.sqrt(h ** 2 + (6378 + 300 + h / param) ** 2)

            # Calculate the conditions for the integral calculations based on the radius condition
            # ,as given in the range of the integration in Eq. 27 in SI.
            condition = param_data['radius'] <= radius_condition
            condition2 = param_data['radius'] <= radius_David

            # Calculate the integrals based on the conditions defined above
            integral_1 = param_data.loc[condition & condition2, 'infWRA'].cumsum() * 100
            integral_2 = (param_data.loc[~condition & condition2, 'infWRA'] * -1).cumsum() * 100

            if not len(integral_2) == 0:
                result = pd.concat([integral_1, integral_1.iloc[-1] + integral_2]).sort_index()
                param_data.loc[:, 'result'] = result
                # print(param, "solution")
            else:
                print("radius_David:", radius_David, "No solution")
                min_x.append(radius_David - 6378)


        result = pd.concat([integral_1, integral_1.iloc[-1] + integral_2]).sort_index()
        param_data.loc[:, 'result'] = result
        label_str = r'$\frac{rk^\phi}{k^r} (\approx b_{ph})$' + f'= {param}'
        ax.plot(param_data['radius_minus_r_sender'], param_data['result'], label=label_str, color=color)
        param_data = param_data[param_data['radius_minus_r_sender'] >= rmin]
        max_x.append(param_data['radius_minus_r_sender'].max())
        min_result.append(param_data['result'].min())
        max_result.append(param_data['result'].max())

    # Set the labels and legend for the plot
    ax.set_xlabel('Altitude of David (km)', fontsize=18)
    ax.set_ylabel('Relative Wigner rotation angle (degree)', fontsize=18)
    ax.tick_params(axis='both', which='major', labelsize=15)
    legend = ax.legend(ncol=2, loc='best', fontsize=13)
    legend.get_title().set_fontsize(20)

    # Set up the formatter for the x and y axes to make the plot more readable
    for axis, setter in [(ax.xaxis, 'xaxis'), (ax.yaxis, 'yaxis')]:
        formatter = ScalarFormatter(useMathText=True)
        formatter.set_scientific(True)
        formatter.set_powerlimits((-3, 3))
        axis.set_major_formatter(formatter)
        getattr(axis, 'offsetText').set_fontsize(14)

    # Set the limits for the x and y axes based on the data
    ax.set_xlim(max(min_x), max(max_x))
    ax.set_ylim(min(min_result), 0)

    # Adjust the layout to make the plot more aesthetically pleasing
    plt.tight_layout()

    plt.savefig('Results/relative WRA.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/relative WRA.svg', dpi=300, bbox_inches='tight')
    plt.show()

def plot_wigner_rotation_BH(file_path1, file_path2):
    fig, ax = plt.subplots(figsize=(8, 5))

    plot_wigner_rotation(file_path1, Earth=False, ax=ax)
    plot_wigner_rotation(file_path2, Earth=False,negative_params=True, ax=ax)
    plt.savefig('Results/WRA_BH' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/WRA_BH' + '.svg', dpi=300, bbox_inches='tight')
    plt.show()
def plot_polar_BH(file_path1, file_path2):
    fig, ax = plt.subplots(figsize=(8, 5))
    plot_wigner_rotation(file_path1, Earth=False, ax=ax)
    plot_wigner_rotation(file_path2, Earth=False, negative_params=True, ax=ax)
    plt.savefig('Results/polar_BH' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/polar_BH' + '.svg', dpi=300, bbox_inches='tight')
    plt.show()
def plot_wigner_rotation_BH_for_SI(file_path1, file_path2, file_path3, file_path4):
    data1 = WRA_function(file_path1, Earth=False)
    data2 = WRA_function(file_path2, Earth=False, negative_params=True)

    data_for_zeroJ1 = WRA_function(file_path3, Earth=False)
    data_for_zeroJ2 = WRA_function(file_path4, Earth=False, negative_params=True)

    data_combined = data1.append(data2)
    data_for_zeroJ_combined = data_for_zeroJ1.append(data_for_zeroJ2)

    parameters = data_combined['parameter'].unique()
    positive_parameters = [param for param in parameters if param >= 0]
    negative_parameters = [param for param in parameters if param < 0]

    colors_positive = cm.Reds(np.linspace(0.3, 1, len(positive_parameters)))
    colors_negative = cm.Blues(np.linspace(0.3, 1, len(negative_parameters)))

    param_colors = {param: color for param, color in zip(positive_parameters, colors_positive)}
    param_colors.update({param: color for param, color in zip(negative_parameters, colors_negative)})

    plt.style.use('seaborn-notebook')
    fig, ax = plt.subplots(figsize=(8, 5))

    # Define xmin and xmax
    xmax = min(
        data_combined.groupby('parameter')['radius_minus_r_sender'].max().min(),
        data_for_zeroJ_combined.groupby('parameter')['radius_minus_r_sender'].max().min()
    )

    # Sort the parameters list in descending order
    sorted_parameters = sorted(parameters, reverse=True)

    # Calculate the differences
    data_diff = data_combined.copy()
    data_diff['integrated_infWRA'] = data_combined['integrated_infWRA'] - data_for_zeroJ_combined['integrated_infWRA']

    for param in sorted_parameters:
        color = param_colors[param]
        param_data = data_diff[data_diff['parameter'] == param]
        ax.plot(param_data['radius_minus_r_sender'], param_data['integrated_infWRA'], label=param,
                      color=color)

    ax.set_xlabel('Distance from 4.5$r_{s}$ (km)', fontsize=10, labelpad=20)
    ax.set_ylabel(r'$\Delta$WRA b.t.w $J=0$ and $J\ne0$', fontsize=10)
    ax.set_xlim(0, xmax)  # Update x-axis range for the inset plot

    # Set the title fontsize
    # legend.get_title().set_fontsize(20)
    formatter = ScalarFormatter(useMathText=True)
    formatter.set_scientific(True)
    formatter.set_powerlimits((-3, 3))
    ax.xaxis.set_major_formatter(formatter)
    # ax_inset.xaxis.set_major_formatter(formatter)
    ax.xaxis.offsetText.set_fontsize(14)
    ax.yaxis.offsetText.set_fontsize(14)

    formatter_y = ScalarFormatter(useMathText=True)
    formatter_y.set_scientific(True)
    formatter_y.set_powerlimits((0, 1))
    # ax_inset.yaxis.set_major_formatter(formatter_y)
    ax.set_xlabel('Distance from 4.5$r_{s}$ (km)', fontsize=18)
    ax.set_ylabel('Wigner rotation angle difference (degree)', fontsize=18)
    ax.tick_params(axis='both', which='major', labelsize=15)

    plt.tight_layout()

    plt.savefig('Results/WRA_difference_BH_spinning_angular_mommentum' + '.png', dpi=300, bbox_inches='tight')
    plt.savefig('Results/WRA_difference_BH_spinning_angular_mommentum' + '.svg', dpi=300, bbox_inches='tight')
    plt.show()

plot_wigner_rotation_for_both_positive_and_negative_impact_parameter(file_path_rearth, file_path_rearth_m)
plot_wigner_rotation_difference(file_path, file_path_m)
plot_wigner_rotation_BH(file_path_BH, file_path_BH_m)
plot_wigner_rotation_BH_for_SI(file_path_BH, file_path_BH_m, file_path_BH_zeroJ, file_path_BH_m_zeroJ)
plot_cofactor_for_positive_impact_parameter(file_path_cofactor_rearth)
plot_cofactor_for_negative_impact_parameter(file_path_cofactor_rearth_m)

plot_cofactor_for_positive_impact_parameter_BH(file_path_cofactor_BH)
plot_cofactor_for_negative_impact_parameter_BH(file_path_cofactor_BH_m)

plot_polar_BH(file_path_BH_zeroLz, file_path_BH_zeroLz_m)